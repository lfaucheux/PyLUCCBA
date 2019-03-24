# -*- coding: utf8 -*-
from __future__ import print_function

__authors__ = [
    "Marion Dupoux <marion.dupoux@gu.se>",
    "Laurent Faucheux <laurent.faucheux@hotmail.fr>"
]

__all__ = [
    'Dashboard',
    'DataReader',
    'InMindWithCorrespondingUnit',
    'cast',
    'change_rate_extractor',
    'csv_dicter',
    'dict_time_serie_as_row_array',
    'get_file_as_list_of_lines',
    'plt',
    'poler',
    'redim_row_array',
    'save_dir_and_file_name',
    'solver_msgr',
    'solver_ND',
    'taber',
    'txt_dicter',
    'xlsx_file_writer',
]

from matplotlib.font_manager import FontProperties
from scipy.optimize import fsolve
import matplotlib.pyplot as plt
import functools as ft
import openpyxl as xl
import shutil as sh
import numpy as np
import sys
import os
import re

VERBOSE_DTESTS = False
OS_SEP = os.sep
__name__eq__main__ = __name__ == '__main__'

##******************************************
##    ┌─┐┌─┐┌─┐┌┬┐
##    │  ├─┤└─┐ │ 
##    └─┘┴ ┴└─┘ ┴
def cast(s):
    """ Function which clarifies the implicit type of
    strings, a priori coming from csv-like files.


    Example
    -------
    >>> cast('1')
    1
    >>> cast('1.')
    1.0
    >>> cast('1E+0')
    1.0
    >>> cast('one')
    'one'
    """
    s = str(s)
    try:
        float(s)
        if '.' in s or 'E' in s:
            return float(s)
        else:
            return int(s)
    except:
        return s

##******************************************
##    ┌─┐┌─┐┌┬┐   ┌─┐┬┬  ┌─┐    ┌─┐┌─┐    ┬  ┬┌─┐┌┬┐   ┌─┐┌─┐  ┬  ┬┌┐┌┌─┐┌─┐
##    │ ┬├┤  │    ├┤ ││  ├┤     ├─┤└─┐    │  │└─┐ │    │ │├┤   │  ││││├┤ └─┐
##    └─┘└─┘ ┴────└  ┴┴─┘└─┘────┴ ┴└─┘────┴─┘┴└─┘ ┴────└─┘└────┴─┘┴┘└┘└─┘└─┘
def get_file_as_list_of_lines(fname):
    """ Function which gets the content of a given file as a list of its lines.

    Example
    -------
    >>> get_file_as_list_of_lines(
    ...     fname = os.path.join(
    ...         'resources', 'yields', 'Output', 'ETH_yields_FR.txt'
    ...     ),
    ... )[:2]
    ['O:unit:tonne[Output]/tonne[Output]', 'O:yrb:2007']
    """
    with open(fname, 'r') as f:
        return [
            l.replace('\n','').replace('\r','')
            for l in f.readlines()
        ]

##******************************************
##    ┌─┐┌─┐┬  ┬   ┬┌┐┌╔╦╗┬┌┐┌┌┬┐
##    │  └─┐└┐┌┘   ││││║║║││││ ││
##    └─┘└─┘ └┘────┴┘└┘╩ ╩┴┘└┘─┴┘
def csv_dicter(wkey, fname, pop):
    """ Function which gets the content of a given file as dict whose file-lines
    are identifyed via `wkey`. NB : fname a full path and name of the given file.

    Testing/Example
    ---------------
    >>> csv_as_dict = csv_dicter(
    ...     wkey  = 'year',
    ...     fname = os.path.join(
    ...         'resources', 'prices', 'Exput', 'CO2_prices_FR.csv'
    ...     ),
    ...     pop   = True,
    ... )
    >>> csv_as_dict[2050]['O']
    87
    >>> csv_as_dict[2050]['SPC2009']
    219.1123143
    """
    dico = {}
    l    = get_file_as_list_of_lines(fname)
    keys = l[0].split(';')

    for i_l, line in enumerate(l[1:]):
        values    = [cast(v) for v in line.split(';')]
        dic       = dict(zip(keys, values))
        wid       = dic[wkey]
        wid       = wid.lower() if isinstance(wid, str) else wid
        dico[wid] = dic
        if pop:
            dico[wid].pop(wkey)

    return dico

##******************************************
##    ┌┬┐─┐ ┬┌┬┐   ┬┌┐┌╔╦╗┬┌┐┌┌┬┐
##     │ ┌┴┬┘ │    ││││║║║││││ ││
##     ┴ ┴ └─ ┴────┴┘└┘╩ ╩┴┘└┘─┴┘
def txt_dicter(fname):
    """ Function which gets the content of a given file as dict, but with no
    file-lines identifiers.
      NB1 : fname a full path and name of the given file
      NB2 : line identifier is supposed to be in the first place,
            i.e. implicit column of the file.

    Testing/Example
    ---------------
    >>> txt_as_dict = txt_dicter(
    ...     fname = os.path.join(
    ...         'resources', 'prices', 'Exput', 'Exput.txt'
    ...     ),
    ... )
    
    >>> scenario_name = 'spc2009'
    >>> sorted(txt_as_dict[scenario_name].items())
    [('unit', 'EUR/tonne'), ('yrb', 2008)]
    
    >>> scenario_name = 'weo2015-450s'
    >>> sorted(txt_as_dict[scenario_name].items())
    [('unit', 'USD/tonne'), ('yrb', 2014)]
    """
    splitted_fName   = fname.split(OS_SEP)[:-1]
    containing_foler = splitted_fName[-1]
    filepath         = OS_SEP.join(fname.split(OS_SEP)[:-1])
    txt_fName_a      = fname.replace('.csv', '.txt')
    txt_fName_b      = os.path.join(filepath, '%s.txt'%containing_foler)
    if os.path.isfile(txt_fName_a):
        txt_fName = txt_fName_a
    elif os.path.isfile(txt_fName_b):
        txt_fName = txt_fName_b
    else:
        raise type(
            'NoDescripterError',
            (BaseException,), {}
        )(
            '[!!!] No txt file descripter present [!!!] '
            '\n\t Look what the problem is @ \n\t\t %s '%filepath
        )

    l    = get_file_as_list_of_lines(txt_fName)
    dico = {}
    for element in l:
        key, descrip_key, descrip_val = element.split(':')[:3]
        key = key.lower()
        if key not in dico:
            dico[key] = {}
        dico[key][descrip_key] = cast(descrip_val)
    return dico

##******************************************
##    ┌┬┐┌─┐┌┐ ┌─┐┬─┐
##     │ ├─┤├┴┐├┤ ├┬┘
##     ┴ ┴ ┴└─┘└─┘┴└─
def taber(msg, size):
    """ Function which tabulates strings.

    Example
    -------
    >>> taber(msg='Example string', size=42)
    'Example string                            '
    """
    try:
        return msg + (size - len(msg))*' '
    except:
        return str(msg) + (size - len(str(msg)))*' '

##******************************************
##    ┌─┐┌─┐┬ ┬  ┬┌─┐┬─┐    ┌┬┐┌─┐┌─┐┬─┐
##    └─┐│ ││ └┐┌┘├┤ ├┬┘    │││└─┐│ ┬├┬┘
##    └─┘└─┘┴─┘└┘ └─┘┴└─────┴ ┴└─┘└─┘┴└─
def solver_msgr(t, c, s, v, e):
    """ Enriching wrapper of the function `taber` defined for solver's
    messages

    Testing/Example
    ---------------
    >>> msg = solver_msgr(
    ...     t = 'Title of what is going on and whose',
    ...     c = 'The solution (may not have) converged.',
    ...     s = 0.,
    ...     v = '[???]',
    ...     e = 1e-15
    ... )
    >>> print(msg)
    ---- Title of what is going on and whose sol=0.0
    ---- [???]The solution (may not have) converged.[1.000000e-15][???]
    """
    sol_section = taber(
        '---- %s sol=%s'%(t[:70], s), 0
    )
    com_section = '\n---- {0}{1}[{2}]{0}'.format(
        v, c.replace(' \n ', ''), '%e'%e
    )
    return  sol_section + com_section

##******************************************
##    ┌─┐┌─┐┬ ┬  ┬┌─┐┬─┐    ╔╗╔╔╦╗
##    └─┐│ ││ └┐┌┘├┤ ├┬┘    ║║║ ║║
##    └─┘└─┘┴─┘└┘ └─┘┴└─────╝╚╝═╩╝
def solver_ND(display, title, func, _Z_, bforce, *args, **kwargs):
    """ N dimensions solver.

    Example
    -------
    >>> first_guess = [1.7]
    >>> solver_ND(
    ...     display = True,
    ...     title   = 'Golden ratio',
    ...     func    = lambda x: x**2 - x - 1,
    ...     _Z_     = first_guess,
    ...     bforce  = False,
    ... )
    ---- Golden ratio sol=[1.61803399]
    ---- [***]The solution converged.[0.000000e+00][***]
    array([1.61803399])

    
    >>> first_guess = [3.2]
    >>> solver_ND(
    ...     display = True,
    ...     title   = 'Pi',
    ...     func    = lambda x: np.sin(x),
    ...     _Z_     = first_guess,
    ...     bforce  = False,
    ... )
    ---- Pi sol=[3.14159265]
    ---- [***]The solution converged.[1.224647e-16][***]
    array([3.14159265])

    >>> solver_ND(
    ...     display = True,
    ...     title   = 'Phi and Pi',
    ...     func    = lambda xy: [xy[0]**2 - xy[0] - 1, np.sin(xy[1])],
    ...     _Z_     = [1.7, 3.2],
    ...     bforce  = False,
    ... )
    ---- Phi and Pi sol=[1.61803399 3.14159265]
    ---- [***]The solution converged.[2.048476e-13][***]
    array([1.61803399, 3.14159265])

    """
    ans = fsolve(
        func, _Z_, full_output=True, *args
    )
    _S_, kwinfo, info, comm = ans[:4]

    if np.sum(np.abs(kwinfo['fvec']))>1e-10:
        info = 4
        _S_ *= .9

    if info not in [1, 5]:
        return solver_ND(
            display, title, func, _S_, bforce, *args, **kwargs
        )
    elif info in [1]:
        if display:
            msg = solver_msgr(
                title, comm, _S_, '[***]',
                np.sum(np.abs(func(_S_, *args)))
            )
            print(msg)
    elif info in [5]:
        if bforce:
            if func(_S_)!=func(_Z_):
                return solver_ND(
                    display, title, func, _S_, bforce, *args, **kwargs
                )            
        msg = solver_msgr(
            title, comm, _S_, '[!!!]',
            np.sum(np.abs(func(_S_, *args)))
        )
        if display:
            print(msg)
        if kwargs.get('msg', False):
            return msg
    return _S_

##******************************************
##    ┌─┐┬ ┬┌─┐┌┐┌┌─┐┌─┐    ┬─┐┌─┐┌┬┐┌─┐    ┌─┐─┐ ┬┌┬┐┬─┐┌─┐┌─┐┌┬┐┌─┐┬─┐
##    │  ├─┤├─┤││││ ┬├┤     ├┬┘├─┤ │ ├┤     ├┤ ┌┴┬┘ │ ├┬┘├─┤│   │ │ │├┬┘
##    └─┘┴ ┴┴ ┴┘└┘└─┘└─┘────┴└─┴ ┴ ┴ └─┘────└─┘┴ └─ ┴ ┴└─┴ ┴└─┘ ┴ └─┘┴└─
def change_rate_extractor(change_rates, initial_currency, final_currency):
    """ Function which tests directions of exchange factors and returns the
    appropriate conversion factor.

    Example
    -------
    >>> change_rate_extractor(
    ...     change_rates     = {'EUR/USD': .8771929824561404},
    ...     initial_currency = 'EUR',
    ...     final_currency   = 'USD',
    ... )
    1.14
    
    >>> change_rate_extractor(
    ...     change_rates     = {'USD/EUR': 1.14},
    ...     initial_currency = 'EUR',
    ...     final_currency   = 'USD',
    ... )
    1.14

    """
    ACR_1 = '%s/%s'%(
        initial_currency, final_currency
    )
    ACR_2 = '%s/%s'%(
        final_currency, initial_currency
    )
    if ACR_1 in change_rates:
        return pow(change_rates[ACR_1], -1.)
    if ACR_2 in change_rates:
        return change_rates[ACR_2]

##******************************************
##    ┌┬┐┬┌─┐┌┬┐  ┌┬┐┬┌┬┐┌─┐    ┌─┐┌─┐┬─┐┬┌─┐    ┌─┐┌─┐    ┬─┐┌─┐┬ ┬    ┌─┐┬─┐┬─┐┌─┐┬ ┬
##     ││││   │    │ ││││├┤     └─┐├┤ ├┬┘│├┤     ├─┤└─┐    ├┬┘│ ││││    ├─┤├┬┘├┬┘├─┤└┬┘
##    ─┴┘┴└─┘ ┴────┴ ┴┴ ┴└─┘────└─┘└─┘┴└─┴└─┘────┴ ┴└─┘────┴└─└─┘└┴┘────┴ ┴┴└─┴└─┴ ┴ ┴ 
def dict_time_serie_as_row_array(time_serie):
    """ Function which turns a dict into a numpy.array that is sorted according
    to the keys of the inputed dict.

    Example
    -------
    >>> dict_time_serie_as_row_array(
    ...     {2020:1., 2015:1.1, 2008:1.2, 2025:1.3},
    ... )
    array([[1.2, 1.1, 1. , 1.3]])
    """
    return np.array([time_serie[t] for t in sorted(time_serie.keys())])[None, :]

##******************************************
##    ┬─┐┌─┐┌┬┐┬┌┬┐    ┬─┐┌─┐┬ ┬    ┌─┐┬─┐┬─┐┌─┐┬ ┬
##    ├┬┘├┤  ││││││    ├┬┘│ ││││    ├─┤├┬┘├┬┘├─┤└┬┘
##    ┴└─└─┘─┴┘┴┴ ┴────┴└─└─┘└┴┘────┴ ┴┴└─┴└─┴ ┴ ┴ 
def redim_row_array(T_array, T_, T0_):
    """ Function which reshapes T_array by adding it 0 if `T<T0`.

    Example
    -------
    >>> a = np.ones((1, 4))

    >>> redim_row_array(
    ...     T_array = a,
    ...     T_      = 2010,
    ...     T0_     = 2012,
    ... )
    array([[1., 1., 1., 1., 0., 0.]])

    >>> redim_row_array(
    ...     T_array = a,
    ...     T_      = 2012,
    ...     T0_     = 2010,
    ... )
    array([[1., 1., 1., 1.]])
    """
    T, T0 = int(T_), int(T0_)
    return np.hstack((
        T_array,
        np.zeros((1, T0 - T)),
    )) if T0>T else T_array[:, :T0]

##******************************************
##    ┌─┐┌─┐┬  ┌─┐┬─┐
##    ├─┘│ ││  ├┤ ├┬┘
##    ┴  └─┘┴─┘└─┘┴└─
def poler(sparse_trajectory_as_dict, repeated_pattern_polation, **kwargs):
    """ Function which retropolates/interpolates/extrapolates values.

    Example
    -------
    >>> sptraj = {
    ...     2010: 1,
    ...     2012: 1.2,
    ...     2014: 1.1,
    ... }
    >>> cst_dense_traj = poler(
    ...     sparse_trajectory_as_dict = sptraj,
    ...     repeated_pattern_polation = False,
    ...     y0 = 2008, yT = 2016
    ... )
    >>> sorted(cst_dense_traj.items())
    [(2009, 0.9128709291752769), (2010, 1.0000000000000002), (2011, 1.0954451150103321), (2012, 1.2), (2013, 1.148912529307606), (2014, 1.1), (2015, 1.1)]
    
    >>> rep_dense_traj = poler(
    ...     sparse_trajectory_as_dict = sptraj,
    ...     repeated_pattern_polation = True,
    ...     y0 = 2008, yT = 2016
    ... )
    >>> sorted(rep_dense_traj.items())
    [(2009, 0.9128709291752769), (2010, 1.0000000000000002), (2011, 1.0954451150103321), (2012, 1.2), (2013, 1.148912529307606), (2014, 1.1), (2015, 1.2049896265113655)]

    """
    #------------------< horizon >
    year_base0, final_year    = kwargs.get('y0', 1950), kwargs.get('yT', 2400)
    sparse_trajectory_as_dict = {
        y : sparse_trajectory_as_dict.get(y)
        for y in range(year_base0, final_year, 1)
    }
    #------------------< core job >
    year_base, varTraj = year_base0 + 0, {}
    year_polate        = sorted(sparse_trajectory_as_dict.keys())
    min_years          = []
    antmin_years       = []
    polated_values     = {}
    rate               = .0
    from_the_left      = True
    #----------< left values >
    while True:
        #--< get minimum year not in memory still >
        try:
            min_year = min(
                filter(
                    lambda y: y not in min_years and sparse_trajectory_as_dict[y],
                    year_polate
                )
            )
            min_years.append(min_year)
        except Exception as exc:
            break
        
        #--< get anteminimum year not in memory still >
        try:
            antmin_year = min(
                filter(
                    lambda y: y not in min_years and sparse_trajectory_as_dict[y],
                    year_polate
                )
            )
            antmin_years.append(antmin_year)
            rate = -1. + pow(
                1.*sparse_trajectory_as_dict[antmin_year]\
                /sparse_trajectory_as_dict[min_year],
                1./(antmin_year - min_year)
            )
        except Exception as exc:
            #raw_input('II: %s'%E)
            #< maintain just previous years trunk variation rate >
            from_the_left = False
            antmin_year   = year_polate[-1]

        #--< iteratively fill concerned years >
        i_aD,i_aG = 0,0
        for year in year_polate:
            if year_base < year <= antmin_year:
                if from_the_left:
                    i_aG    += 1
                    ref_year = antmin_year
                    #< exponential variation >
                    varTraj[len(varTraj.keys())] = 1.+ rate
                    polated_values[year] = 1/pow(
                        1. + rate, float(ref_year - year)
                    )*sparse_trajectory_as_dict[ref_year]
                else:
                    i_aD += 1
                    #< logistic variation (with the previous years trunk variation rate) >
                    if i_aD==1:
                        #< years-invariant conditions >
                        ref_year = min_year
                        v        = float(sparse_trajectory_as_dict[ref_year])
                        v0       = v
                        Kmax     = 1.5*v
                        Kmin     = 0.75*v
                    N = ref_year-year # <0
                    r = rate
                    if repeated_pattern_polation and len(varTraj):
                        try:
                            B  = max(
                            Kmin,
                            min(
                                Kmax,
                                v*varTraj[(i_aD-1)%len(varTraj.keys())]
                            )
                        )
                        except:
                            B = v
                    else:   B = v
                    polated_values[year] = B
                    
                    #< prepare next initial conditions >
                    ref_year = year
                    v        = B
            if year > antmin_year:
                year_base = 1*antmin_year
                break
    #---------------------< control for empty dictionnary and return >
    return polated_values if len(polated_values)\
           else {year : 1 for year in year_polate}

##******************************************
##    ╔═╗┌─┐┌─┐┬ ┬┌─┐
##    ║  ├─┤│  ├─┤├┤ 
##    ╚═╝┴ ┴└─┘┴ ┴└─┘
class Cache(object):

    def __init__(self, *args, **kwargs):
        """ Homemade cache class which aims at being inherited """
        self._cache  = {}
        self.verbose = kwargs.get('verbose', False)

    def _clear_cache(self):
        """
        Testing
        -------
        >>> c = Cache()
        >>> c._cache['key'] = 'value'
        >>> c._clear_cache()
        (1, 0)
        """
        l0 = len(self._cache)
        self._cache.clear()
        return (l0, len(self._cache))

    def verboser(self, _cache, _key_):
        """ Function which traces calculations and returns infos about
        values beeing computed, such as type, dimension (shape or len),
        variable name.

        Testing/Example
        ---------------
        >>> o = Cache(verbose=True)

        >>> o._cache['an_array'] = np.ones((1, 3))
        >>> o.verboser(o._cache, 'an_array')
        ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        key   : an_array
        type  : ndarray
        shape : (1, 3)
        sum   : 3.0
        len   : 1
        minkey: 1.0
        maxkey: 1.0
        value : [[1. 1. 1.]]

        >>> o._cache['a_float'] = 1.
        >>> o.verboser(o._cache, 'a_float')
        ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        key   : a_float
        type  : float
        value : 1.0

        >>> o._cache['a_list'] = [1., 1., 1.]
        >>> o.verboser(o._cache, 'a_list')
        ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        key   : a_list
        type  : list
        sum   : 3.0
        len   : 3
        minkey: 1.0
        maxkey: 1.0
        value : [1.0, 1.0, 1.0]

        >>> o._cache['a_dict'] = {2000: 1., 2100:0.}
        >>> o.verboser(o._cache, 'a_dict')
        ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        key   : a_dict
        type  : dict
        minkey: 2000
        maxkey: 2100
        value : [(2000, 1.0), (2100, 0.0)]

        >>> o._cache['endogenizing'] = True
        >>> o.verboser(o._cache, 'a_list')
        """
        if self.verbose and not _cache.get('endogenizing', False):
            value = _cache[_key_]
            print(50*'~')
            print('key   :', _key_)

            print('type  :', type(value).__name__)

            if isinstance(value, np.ndarray):
                print('shape :', value.shape)

            if isinstance(value, np.ndarray) or isinstance(value, list):
                print('sum   :', np.sum(value))
                print('len   :', len(value))
                print('minkey:', np.min(value))
                print('maxkey:', np.max(value))
            
            if isinstance(value, dict):
                v_keys = value.keys()
                print('minkey:', min(v_keys))
                print('maxkey:', max(v_keys))
                print('value :', sorted(value.items()))
            else:                
                print('value :', value)

    @classmethod
    def _property(cls, meth):
        """ Memoizes outcomes of the so-decorated attribute. The name of the
        attribute is used as identifier.

        Testing/Example
        ---------------
        >>> import random as rd
        >>> class_ = type(
        ...     'class_', 
        ...     (Cache, ), 
        ...     {
        ...         'suffix': ' plus an id',
        ...         'attr'  : Cache._property(
        ...             meth = lambda cls: rd.random(),
        ...         )
        ...     }, 
        ... )
        >>> o = class_()
        >>> o.attr == o.attr
        True
        """
        @property
        @ft.wraps(meth)
        def __property(cls, *args, **kwargs):
            meth_name = meth.__name__
            if meth_name not in cls._cache:
                cls._cache[meth_name] = meth(cls, *args, **kwargs)
                cls.verboser(cls._cache, meth_name)
            return cls._cache[meth_name]
        return __property

##******************************************
##    ╦┌┐┌╔╦╗┬┌┐┌┌┬┐╦ ╦┬┌┬┐┬ ┬╔═╗┌─┐┬─┐┬─┐┌─┐┌─┐┌─┐┌─┐┌┐┌┌┬┐┬┌┐┌┌─┐╦ ╦┌┐┌┬┌┬┐
##    ║│││║║║││││ ││║║║│ │ ├─┤║  │ │├┬┘├┬┘├┤ └─┐├─┘│ ││││ │││││││ ┬║ ║││││ │ 
##    ╩┘└┘╩ ╩┴┘└┘─┴┘╚╩╝┴ ┴ ┴ ┴╚═╝└─┘┴└─┴└─└─┘└─┘┴  └─┘┘└┘─┴┘┴┘└┘└─┘╚═╝┘└┘┴ ┴ 
class InMindWithCorrespondingUnit(Cache):
    """ Class which gets file values and units correspondig to its keys
      NB : A convention adopted in this project is :
      * csv files for numerical data
      * txt files for information about numerical data."""

    def __init__(self, wkey, fname, pop=True):
        super(InMindWithCorrespondingUnit, self).__init__()
        self.wkey  = wkey
        self.fname = fname
        self.pop   = pop

    @Cache._property
    def keys_and_values(self):
        """
        Testing/Example
        ---------------
        >>> o = InMindWithCorrespondingUnit(
        ...     wkey  = 'year',
        ...     fname = os.path.join(
        ...         'resources', 'yields', 'Output', 'ETH_yields_FR.csv'
        ...     ),
        ... )
        >>> o.keys_and_values[2040]['DEBUG']
        1
        """
        return csv_dicter(self.wkey, self.fname, pop=self.pop)

    @Cache._property
    def keys_and_infos(self):
        """
        Testing/Example
        ---------------
        >>> o = InMindWithCorrespondingUnit(
        ...     wkey  = '',
        ...     fname = os.path.join(
        ...         'resources', 'prices', 'Exput', 'Exput.txt'
        ...     )
        ... )
        >>> kinf = o.keys_and_infos['weo2015-nps']
        >>> kinf['yrb']
        2014
        >>> kinf['unit']
        'USD/tonne'
        """
        return txt_dicter(self.fname)

    @Cache._property
    def values_and_infos_per_key(self):
        """
        Testing/Example
        ---------------
        >>> o = InMindWithCorrespondingUnit(
        ...     wkey  = '',
        ...     fname = os.path.join(
        ...         'resources', 'dluc', 'CS_yields_FR.csv'
        ...     ),
        ... )
        >>> soc = o.values_and_infos_per_key['soc']
        >>> soc['infos']
        {'unit': 'Tonne/ha'}
        >>> sorted(
        ...     soc['values'].items()
        ... )[:3]
        [('ANNUAL CROPLAND', 64.73754052), ('DEBUG', 51.33333333), ('DEGRADED GRASSLAND', 49.93333333)]
        """
        return {
            key: {
                'values':self.keys_and_values.get(key),
                'infos' :self.keys_and_infos.get(key),
            } for key in self.keys_and_values.keys()
        }

##******************************************
##    ┌─┐┌─┐┬  ┬┌─┐    ┌┬┐┬┬─┐    ┌─┐┌┐┌┌┬┐    ┌─┐┬┬  ┌─┐    ┌┐┌┌─┐┌┬┐┌─┐
##    └─┐├─┤└┐┌┘├┤      │││├┬┘    ├─┤│││ ││    ├┤ ││  ├┤     │││├─┤│││├┤ 
##    └─┘┴ ┴ └┘ └─┘─────┴┘┴┴└─────┴ ┴┘└┘─┴┘────└  ┴┴─┘└─┘────┘└┘┴ ┴┴ ┴└─┘
def save_dir_and_file_name(save_dir='', file_name='', save=True):
    """
    Function which checks filenames and creates non-existing directories
    on the fly.
    
    Testing/Example
    ---------------
    >>> save_dir_and_file_name(
    ...     save_dir  = '',
    ...     file_name = 'textfile.txt'
    ... )
    ('', 'textfile.txt')
    
    >>> save_dir_and_file_name(
    ...     save_dir  = '',
    ...     file_name = ''
    ... )
    ('', 'no_name')
    """
    if len(save_dir):
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        save_dir += OS_SEP
    file_name = 'no_name' if save and not len(file_name) else file_name
    return save_dir, file_name

##******************************************
##    ─┐ ┬┬  ┌─┐─┐ ┬    ┌─┐┬┬  ┌─┐    ┬ ┬┬─┐┬┌┬┐┌─┐┬─┐
##    ┌┴┬┘│  └─┐┌┴┬┘    ├┤ ││  ├┤     │││├┬┘│ │ ├┤ ├┬┘
##    ┴ └─┴─┘└─┘┴ └─────└  ┴┴─┘└─┘────└┴┘┴└─┴ ┴ └─┘┴└─
try:
    op_get_column_letter = xl.utils.get_column_letter
except Exception as exc:
    print(exc)
    op_get_column_letter = xl.cell.get_column_letter
def xlsx_file_writer(listed_content, save_dir='', file_name=''):
    """ Function which writes lists' contents in xlsx files by assuming
    that `listed_content` is two-dimensional and using elements indexes
    to locate them on sheets.

    Testing/Example
    ---------------
    >>> xlsx_file_writer(
    ...     listed_content = [['A1', 'B1'], ['A2', 'B2']],
    ... )
    True
    >>> os.remove('no_name.xlsx')
    """
    save_dir, file_name = save_dir_and_file_name(
        save_dir=save_dir,
        file_name=file_name
    )
    wb = xl.Workbook()
    ws = wb.active
    for i_row, row in enumerate(listed_content):
        for i_col, col in enumerate(row):
            ws['%s%s'%(op_get_column_letter(i_col+1), i_row+1)] = col
    wb.save('%s%s.xlsx'%(save_dir, file_name))
    return True

##******************************************
##    ╔╦╗┌─┐┌┬┐┌─┐╦═╗┌─┐┌─┐┌┬┐┌─┐┬─┐
##     ║║├─┤ │ ├─┤╠╦╝├┤ ├─┤ ││├┤ ├┬┘
##    ═╩╝┴ ┴ ┴ ┴ ┴╩╚═└─┘┴ ┴─┴┘└─┘┴└─
class DataReader(Cache):
    """ Class which maps directories and files needed for computations."""
    def __init__(self, **kwargs):
        super(DataReader, self).__init__()
        self.country         = '_%s'%kwargs.get('country', 'FR')[:2].upper()
        self.package_folder  = os.path.dirname(__file__)
        self.local_folder    = os.getcwd()
        self.from_local_data = kwargs.get('from_local_data', False)

    def _folder_copier(self, name='resources'):
        """ Method used to copy the resource folder locally.

        Testing/Example
        ---------------
        >>> foldername = 'resources'
        >>> dr = DataReader(
        ...     from_local_data = True
        ... )
        >>> os.path.exists(
        ...     os.path.join(dr.package_folder, foldername)
        ... )
        True
        >>> tmp_folder = os.path.join(dr.package_folder, '.tmp')
        >>> dr.local_folder = tmp_folder
        >>> dr._folder_copier(
        ...     name = foldername
        ... )
        >>> sorted(dr.resources.keys())
        ['dluc', 'prices', 'yields']
        >>> tmp_exists = lambda : os.path.exists(tmp_folder)
        >>> tmp_exists()
        True
        >>> while tmp_exists():
        ...     sh.rmtree(tmp_folder, ignore_errors=True)
        >>> tmp_exists()
        False
        """
        if self.local_folder != self.package_folder:
            sh.copytree(
                src = os.path.join(self.package_folder, name),
                dst = os.path.join(self.local_folder, name)
            )
            if not __name__eq__main__:
                print("'{}' folder copied to {}".format(
                    name, self.local_folder
                ))

    @Cache._property
    def resources_folder_dir(self):
        """ Memoized directory of the folder that contain the resources data
        used to conduct studies."""
        return os.path.join(
            self.local_folder if self.from_local_data else self.package_folder,
            'resources'
        )

    @property
    def resources(self):
        """ Wraper of `_resources` used to check whether resources data are
        available where they are said to be so.

        Testing
        -------
        >>> dr = DataReader(country = 'fra')
        >>> os.path.exists(
        ...     os.path.join(dr.package_folder, 'resources')
        ... )
        True
        >>> sorted(dr.resources.keys())
        ['dluc', 'prices', 'yields']
        """
        copy_dir = self.resources_folder_dir
        if not os.path.exists(copy_dir):
            raise type(
                'DataFolderError',
                (BaseException,), {}
            )(
                '\n'.join([
                    22*"=",
                    'You have set the parameter `from_local_data` to '
                    "`True` but you don't have such a folder in your "
                    'working directory. Please first consider making '
                    'a local copy of a folder whose structure will be '
                    'interpretable by PyLUCCBA i.e.',
                    '    >>> from PyLUCCBA import data_resources_copier',
                    '    >>> data_resources_copier()',
                    'You may then want to modify data as you like.'
                ])
            )
        return self._resources

    def show_resources_tree(self):
        """ Method that prints the nested structure of the dictionary
        returned by the method `_resources_mapper`.

        Example
        -------
        >> DataReader(country = 'fRANce').show_resources_tree()
        ------------ dluc
              ------ CS_yields
                    via `data['dluc']['CS_yields']`
                     resources\dluc\CS_yields_FR.csv
        ------------ prices
              ------ Exput
                 --- CO2_prices
                    via `data['prices']['Exput']['CO2_prices']`
                     resources\prices\Exput\CO2_prices_FR.csv
        ------------ yields
              ------ Input
                 --- DEBUG_yields
                    via `data['yields']['Input']['DEBUG_yields']`
                     resources\yields\Input\DEBUG_yields_FR.csv
                 --- HAETH_yields
                    via `data['yields']['Input']['HAETH_yields']`
                     resources\yields\Input\HAETH_yields_FR.csv
                 --- MISCANTHUS_yields
                    via `data['yields']['Input']['MISCANTHUS_yields']`
                     resources\yields\Input\MISCANTHUS_yields_FR.csv
                 --- SUGARBEET_yields
                    via `data['yields']['Input']['SUGARBEET_yields']`
                     resources\yields\Input\SUGARBEET_yields_FR.csv
                 --- WHEAT_yields
                    via `data['yields']['Input']['WHEAT_yields']`
                     resources\yields\Input\WHEAT_yields_FR.csv
              ------ Output
                 --- ETH_yields
                    via `data['yields']['Output']['ETH_yields']`
                     resources\yields\Output\ETH_yields_FR.csv
        """
        for key, obj in sorted(self._resources_mapper().items()):
            print(12*'-', key)
            for subkey, subobj in sorted(obj.items()):
                print(6*' ' + 6*'-', subkey)
                if isinstance(subobj, dict):
                    for subsubkey, subsubobj in sorted(subobj.items()):
                        print(9*' ' + 3*'-', subsubkey)
                        datapath = "`data['{}']['{}']['{}']`".format(
                            key, subkey, subsubkey
                        )
                        print(12*' ' + 'via', datapath)
                        print(12*' ', os.path.relpath(subsubobj))
                else:
                    datapath = "`data['{}']['{}']`".format(
                        key, subkey
                    )
                    print(12*' ' + 'via', datapath)
                    print(12*' ', os.path.relpath(subobj))

    def _resources_mapper(self):
        """ Creates a nested dictionary that represents the folder structure
        of resources.

        Testing/Example
        ---------------
        >>> print_ = lambda p:print(os.path.relpath(p))
        
        >>> ress = DataReader(
        ...     country = 'fRANce'
        ... )._resources_mapper()

        >>> print_(ress['dluc']['CS_yields'])
        resources\dluc\CS_yields_FR.csv

        >>> print_(ress['prices']['Exput']['CO2_prices'])
        resources\prices\Exput\CO2_prices_FR.csv
        
        >>> print_(ress['yields']['Input']['DEBUG_yields'])
        resources\yields\Input\DEBUG_yields_FR.csv
        
        >>> print_(ress['yields']['Input']['HAETH_yields'])
        resources\yields\Input\HAETH_yields_FR.csv
        
        >>> print_(ress['yields']['Input']['MISCANTHUS_yields'])
        resources\yields\Input\MISCANTHUS_yields_FR.csv
        
        >>> print_(ress['yields']['Input']['SUGARBEET_yields'])
        resources\yields\Input\SUGARBEET_yields_FR.csv
        
        >>> print_(ress['yields']['Input']['WHEAT_yields'])
        resources\yields\Input\WHEAT_yields_FR.csv
        
        >>> print_(ress['yields']['Output']['ETH_yields'])
        resources\yields\Output\ETH_yields_FR.csv

        """
        dir_    = {}
        rfolder = self.resources_folder_dir
        rootdir = rfolder.rstrip(OS_SEP)
        start   = rfolder.rfind(OS_SEP) + 1
        for path, dirs, files in os.walk(rootdir):
            folders = path[start:].split(OS_SEP)
            subdir  = {
                f.split('.csv')[0].replace(self.country, '') : os.path.join(
                    path, f
                )
                for f in filter(
                    lambda f_: self.country in f_ and f_.endswith('.csv'),
                    dict.fromkeys(files).keys()
                )
            }
            parent              = ft.reduce(dict.get, folders[:-1], dir_)
            parent[folders[-1]] = subdir
        return dir_['resources']

    @Cache._property
    def _resources(self):
        """ Memoizing wrapper of `_resources_mapper`."""
        return self._resources_mapper()

##******************************************
##    ╔╦╗┌─┐┌─┐┬ ┬┌┐ ┌─┐┌─┐┬─┐┌┬┐
##     ║║├─┤└─┐├─┤├┴┐│ │├─┤├┬┘ ││
##    ═╩╝┴ ┴└─┘┴ ┴└─┘└─┘┴ ┴┴└──┴┘
class Dashboard(object):
    def __init__(self, **kws):
        self.canvas         = plt
        self._return_charts = kws.get('return_charts', True)
        self.prop           = FontProperties()
        self.prop.set_size(10)
        self.prop.set_stretch('ultra-expanded')
        self.prop.set_style('oblique')
        self.prop.set_variant('small-caps')
        self.prop.set_weight('book')
        self.prop.set_family('fantasy')

    _mocked_meth = lambda s:print(
        '!!! Your instance has `return_charts` '
        'set to `False`.\nDo `<your_instance>.'
        'return_charts = True` and retry.'
    )
    def plot(self,
            abs_, imas, labels, colors,
            save=False, save_dir='', file_name='', bar=False
        ):
        """ Method which graphs `imas` against `abs_`.

        Testing/Example
        ---------------
        >>> x  = np.arange(10)[None, :]
        >>> x
        array([[0, 1, 2, 3, 4, 5, 6, 7, 8, 9]])
        >>> ys = np.mgrid[0:3:1, 0:10:1][0, :, :]
        >>> ys
        array([[0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               [1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
               [2, 2, 2, 2, 2, 2, 2, 2, 2, 2]])
        >>> labs = [
        ...     'constant 1',
        ...     'constant 2',
        ...     'constant 3'
        ... ]
        >>> cols = [
        ...     'red', 'blue', 'black'
        ... ]   
        >>> pltobj = Dashboard(return_charts=True).plot(
        ...     abs_   = x.T,
        ...     imas   = ys.T,
        ...     labels = labs,
        ...     colors = cols,
        ... )
        >>> pltobj.show() # doctest: +SKIP
        >>> pltobj.close() 
        >>> pltobj = Dashboard(return_charts=False).plot(
        ...     abs_   = x.T,
        ...     imas   = ys.T,
        ...     labels = labs,
        ...     colors = cols,
        ... )
        >>> pltobj.show()
        !!! Your instance has `return_charts` set to `False`.
        Do `<your_instance>.return_charts = True` and retry.

        """
        if type(abs_).__module__ != np.__name__:
            lenabs_    = len(abs_)
            abs_       = np.array(abs_)
            abs_.shape = (lenabs_, 1)
        imas = imas.T.tolist()

        self.canvas.clf()
        for ima, label,color in zip(imas, labels,colors):
            if type(ima).__module__ != np.__name__:
                lenOrd    = len(ima)
                ima       = np.array(ima)
                ima.shape = (lenOrd,1)
            if bar:
                self.canvas.bar(
                    abs_, ima,
                    align='center',
                    label=label.upper(),
                    color=color,
                    linewidth=0
                )
            else:
                self.canvas.plot(
                    abs_, ima,
                    label=label.upper(),
                    color=color
                )
        self.canvas.legend(
            prop=self.prop,
            loc = 'upper right',
            labelspacing = 0.2,
            bbox_to_anchor =(1.1, 1.)
        )
        save_dir, file_name = save_dir_and_file_name(
            save_dir=save_dir,
            file_name=file_name,
            save=save
        )

        self.canvas.xticks(
            np.arange(min(abs_), max(abs_)+1, 1.0),
            rotation=70
        )

        if save:
            self.canvas.savefig(
                os.path.join(
                    save_dir, '%s.png'%file_name
                ),
                bbox_inches=0, dpi=200
            )
        
        if self._return_charts:
            return self.canvas

        self.canvas.close()
        return type(
            'msger', (object,),{
                'show' : self._mocked_meth,
                'close': self._mocked_meth,
            }
        )

if __name__eq__main__:
    import doctest
    doctest.testmod(verbose=VERBOSE_DTESTS)
